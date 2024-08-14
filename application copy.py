import streamlit as st
import streamlit.components.v1 as components
import pathlib
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
import openpyxl
import pickle
from pathlib import Path
from langchain.document_loaders import YoutubeLoader
from langchain.llms import OpenAI
from langchain.chains.summarize import load_summarize_chain
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os

# OPENAI_API_KEY = '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'

# Functions from the first code block

# Turns pdf into a long string of text
def get_pdf_ortxt_text(docs):
    text = ""
    
    for doc in docs:
        if doc.type.lower() == 'application/pdf':  # Check if the file has a PDF MIME type
            # Process PDF file
            pdf_reader = PdfReader(doc)
            for page in pdf_reader.pages:
                text += page.extract_text()
            
        elif doc.type.lower() == 'text/plain':  # Check if the file has a TXT MIME type
            # Process TXT file
            text += doc.getvalue().decode('utf-8')

        elif doc.type.lower() == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            # Process Excel (XLSX) file
            try:
                excel_workbook = openpyxl.load_workbook(doc)
                for sheet_name in excel_workbook.sheetnames:
                    sheet = excel_workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        text += " ".join(str(cell) for cell in row) + "\n"
            except ImportError:
                print("openpyxl library not available. Please install it.")
            except Exception as e:
                print(f"Error reading Excel file '{doc.name}': {e}")
        
        else:
            print(f"Unsupported file format for '{doc.name}'. Skipping...")
    
    return text

# Functions from the second code block

def create_pdf_from_data(result, output_pdf_filename):
    # Check if the PDF file already exists, and remove it if it does
    if os.path.exists(output_pdf_filename):
        os.remove(output_pdf_filename)

    # Create a PDF canvas
    c = canvas.Canvas(output_pdf_filename, pagesize=letter)

    # Define the initial Y-coordinate for content
    y_position = 700

    # Iterate through each 'item' in the 'result'
    for item in result:
        # Add content to the PDF
        c.drawString(100, y_position, f"Author: {item.metadata['author']}")
        y_position -= 20  # Adjust the Y-coordinate for the next line
        c.drawString(100, y_position, f"Length: {item.metadata['length']} seconds")
        y_position -= 20  # Adjust the Y-coordinate for the next line
        c.drawString(100, y_position, f"Title: {item.metadata['title']}")
        y_position -= 20
        # Add more content as needed
        # Check if 'item.page_content' is not None and add it to the PDF
        if item.page_content:
            y_position -= 20  # Adjust the Y-coordinate for the next line
            content_lines = item.page_content.split('\n')
            for line in content_lines:
                # Ensure that the text wraps within the page width
                if y_position < 50:  # Start a new page if close to the bottom
                    c.showPage()
                    y_position = 700  # Reset Y-coordinate for new page
                c.drawString(100, y_position, line)
                y_position -= 20  # Adjust the Y-coordinate for the next line

        # Move to the next page if there is more content
        c.showPage()

    # Save the PDF file
    c.save()

def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=1000,
        chunk_overlap=200,
        length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks

def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore

def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain

def handle_userinput(user_question):
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']

    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            st.write(bot_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)


def main():
    load_dotenv()
    st.set_page_config(page_title="Multiple Doc Chatbot and Video PDF Creator",
                       page_icon=":books:", layout="wide")
    st.write(css, unsafe_allow_html=True)

    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None

    st.header("Multiple Doc Chatbot and Video PDF Creator :books:")

    # Add a text input widget in the sidebar for user's question
    user_question = st.sidebar.text_input("Ask a question about your documents:")

    # Add a button to trigger user input processing
    if st.sidebar.button("Submit Question"):
        if user_question:
            handle_userinput(user_question)

    # Add a text input widget in the sidebar for video URL
    video_url = st.sidebar.text_input("Enter a YouTube video URL:")

    if video_url:
        if video_url.startswith("https://youtu.be/") or video_url.startswith("https://www.youtube.com/"):
            # Process YouTube video URL
            loader = YoutubeLoader.from_youtube_url(video_url, add_video_info=True)
            result = loader.load()
            output_pdf_filename = 'Video_out.pdf'
            create_pdf_from_data(result, output_pdf_filename)
            st.write(f"Output text PDF file '{output_pdf_filename}' has been created.")
        else:
            st.warning("Please enter a valid YouTube video URL.")

    with st.sidebar:
        st.subheader("Your Documents:")
        docs = st.file_uploader(
            "Upload your Doc (.Pdf,.Txt,.xlsx) here and click on 'Process'", accept_multiple_files=True) # to accept multiple files

        if st.button("Process"):
            with st.spinner("Processing"):
                raw_text = get_pdf_ortxt_text(docs)
                text_chunks = get_text_chunks(raw_text)

            # create vector store
                vectorstore = get_vectorstore(text_chunks)

            # create conversation chain
                st.session_state.conversation = get_conversation_chain(
                vectorstore)


if __name__ == '__main__':
    main()
