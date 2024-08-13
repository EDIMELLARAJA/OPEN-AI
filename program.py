import pathlib
import streamlit as st
from dotenv import load_dotenv
from PyPDF2 import DocumentInformation, PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
import openpyxl 
import pickle
from pathlib import Path


# Turns pdf into a long string of text
def get_pdf_ortxt_text(docs):
    text = ""
    
    for doc in docs:
        if doc.type.lower() == 'application/pdf':  # Check if the file has a PDF MIME type
        # if doc.name.lower().endswith('.pdf'):  # Check if the file has a .pdf extension
            # Process PDF file
            pdf_reader = PdfReader(doc)
            for page in pdf_reader.pages:
                text += page.extract_text()
            
        elif doc.type.lower() == 'text/plain':  # Check if the file has a TXT MIME type
        # elif doc.name.lower().endswith('.txt'):  # Check if the file has a .txt extension
            # Process TXT file
            text += doc.getvalue().decode('utf-8')
            # with open(doc, 'r') as txt_file:
            #     text += txt_file.read()

        elif doc.type.lower() == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            # Process Excel (XLSX) file
            try:
                excel_workbook = openpyxl.load_workbook(doc)
                for sheet_name in excel_workbook.sheetnames:
                    sheet = excel_workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        text += " ".join(str(cell) for cell in row) + "\n"
            except ImportError:
                print("openpyxl library not available. Please install it.")
            except Exception as e:
                print(f"Error reading Excel file '{doc.name}': {e}")
        

        else:
            print(f"Unsupported file format for '{doc.name}'. Skipping...")
    
    return text
        

def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=1000,
        chunk_overlap=200,
        length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks


def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    # embeddings = HuggingFaceInstructEmbeddings(model_name="hkunlp/instructor-xl") 
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore


def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    # llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})

    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain


def handle_userinput(user_question):
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']

    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            st.write(bot_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
            
def main():
    load_dotenv()
    st.set_page_config(page_title="Multiple Doc Chatbot using open AI",
                       page_icon=":books:",layout="wide")
    st.write(css, unsafe_allow_html=True)

    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None

    st.header("Multiple Doc Chatbot using openAI :books:")
    user_question = st.text_input("Ask a question about your documents:")
    if user_question:
        handle_userinput(user_question)

    with st.sidebar:
        st.subheader("Your Documents:")
        docs = st.file_uploader(
            "Upload your Doc (.Pdf,.Txt,.xlsx) here and click on 'Process'", accept_multiple_files=True) # to accept multiple files
    # if txt_file:
    #     txt = get_text_from_txt(txt_file)

    #     chunks = get_text_chunks(txt)
    #     vectorstore = get_vectorstore(chunks)
        if st.button("Process"):
            with st.spinner("Processing"):
            # get pdf text
                raw_text = get_pdf_ortxt_text(docs)

            # get the text chunks
                text_chunks = get_text_chunks(raw_text)

            # create vector store
                vectorstore = get_vectorstore(text_chunks)

            # create conversation chain
                st.session_state.conversation = get_conversation_chain(
                vectorstore)


if __name__ == '__main__':
    main()